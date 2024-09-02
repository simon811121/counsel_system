import os
import csv
import json
from enum import Enum
from runtime_type_checker import check_type
from typing import Any, Dict
from openpyxl import Workbook, load_workbook

class Config:
    def __init__(self,
                 testing: bool):
        self.file_name = "config.json" if testing is False else "config_test.json"
        file_path = gen_file_path(self.file_name, 'data')
        self.file_path = file_path
        self.settings = self._read_settings()

    def _read_settings(self) -> Dict[str, Any]:
        with open(self.file_path, 'r') as file:
            settings = json.load(file)
        return settings

    def get_setting(self, section: str, key: str) -> Any:
        return self.settings.get(section, {}).get(key, None)

    def get_section(self, section: str) -> Dict[str, Any]:
        return self.settings.get(section, {})

    def update_setting(self, section: str, key: str, value: Any):
        if section in self.settings:
            self.settings[section][key] = value
        else:
            self.settings[section] = {key: value}

    def save_settings(self):
        with open(self.file_path, 'w') as file:
            json.dump(self.settings, file, indent=4)

class Account_Info:
    class Info:
        def __init__(self,
                     user_id: int,
                     permission: str,
                     account: str,
                     password: str,
                     name: str,
                     email: str,
                     phone_num: int,
                     register_time: str,
                     last_login_time: str):
            check_type(user_id, int)
            check_type(permission, str)
            check_type(account, str)
            check_type(password, str)
            check_type(name, str)
            check_type(email, str)
            check_type(phone_num, str)
            check_type(register_time, str)
            check_type(last_login_time, str)
            self.allow_perm = {'director', 'psychologist', 'manager', 'guest'}
            if permission not in self.allow_perm:
                raise ValueError("Invalid permission value. Allowed values: 'director', 'psychologist', 'manager', 'guest'")
            self.user_id = user_id
            self.permission = permission
            self.account = account
            self.password = password
            self.name = name
            self.email = email
            self.phone_num = phone_num
            self.register_time = register_time
            self.last_login_time = last_login_time

        def __str__(self):
            return (f"Info(user_id = {self.user_id}\n"
                    f"     permission = {self.permission}\n"
                    f"     account = {self.account}\n"
                    f"     password = {self.password}\n"
                    f"     name = {self.name}\n"
                    f"     email = {self.email}\n"
                    f"     phone_num = {self.phone_num}\n"
                    f"     register_time = {self.register_time}\n"
                    f"     last_login_time = {self.last_login_time})")

        def _equals(self, other):
            if not isinstance(other, Account_Info.Info):
                return False
            return all(getattr(self, attr) == getattr(other, attr) for attr in vars(self))

    def __init__(self,
                 file_name: str = None):
        self.file_name = "account_info.csv" if file_name is None else file_name
        self.acnt_info_fieldnames = ['user_id', 'permission', 'account', 'password', 'name', 'email', 'phone_num', 'register_time', 'last_login_time']
        self.file_path = gen_file_path(self.file_name, 'data')
        self._build_file()
        self.infos = []
        self.glb_cfg = Config(False) if file_name is None else Config(True)
        self.setting = self.glb_cfg.get_section('acnt_info')

    def _build_file(self):
        if not os.path.exists(self.file_path):
            with open(self.file_path, 'w', newline='') as file:
                writer = csv.DictWriter(file, fieldnames=self.acnt_info_fieldnames)
                writer.writeheader()

    def _refresh_file(self): # only for unit-testing
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
            self._build_file()

    def check_account_aval(self,
                           account: str):
        if self.find_acnt_info(account=account) == None:
            return True
        else:
            return False

    def add_acnt_info(self,
                      permission: str,
                      account: str,
                      password: str,
                      name: str,
                      email: str,
                      phone_num: int,
                      register_time: str,
                      last_login_time: str):
        user_id = self.setting['nxt_user_id']
        max_user_id = self.setting['max_user_id']
        if (user_id < max_user_id):
            info = self.Info(user_id, permission, account, password, name, email, phone_num, register_time, last_login_time)
            self.infos.append(info)
            user_id += 1
            self.glb_cfg.update_setting('acnt_info', 'nxt_user_id', user_id)
            self.glb_cfg.save_settings()
            self.file_path = gen_file_path(self.file_name, 'data')
            with open(self.file_path, 'a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow([info.user_id, info.permission, info.account, info.password, info.name, info.email, info.phone_num, info.register_time, info.last_login_time])
            return True
        else:
            return False

    def find_acnt_info(self,
                       user_id: int = None,
                       account: str = None,
                       skip_cache = False):
        if skip_cache == False:
            for infos in self.infos:
                if user_id != None and infos.user_id == user_id:
                    return infos
                elif account != None and infos.account == account:
                    return infos
        with open(self.file_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row in csv_reader:
                # Check if the user_id matches
                infos = None
                if user_id != None and int(row['user_id']) == user_id:
                    row['user_id'] = int(row['user_id']) # read from csv, so type of user_id will be str
                    infos = self.Info(**row)
                elif account != None and row['account'] == account:
                    row['user_id'] = int(row['user_id']) # read from csv, so type of user_id will be str
                    infos = self.Info(**row)
                self.infos.append(infos)
                return infos
            return None

    def update_acnt_info(self,
                         user_id: int = None,
                         account: str = None,
                         password: str = None,
                         name: str = None,
                         email: str = None,
                         phone_num: str = None,
                         last_login_time: str = None):
        info = self.find_acnt_info(user_id=user_id, account=account) # check if info exist
        if (info == None):
            return False

        # gen new from
        update_values = {
            'password': password,
            'name': name,
            'email': email,
            'phone_num': phone_num,
            'last_login_time': last_login_time
        }
        update_values = {key: value for key, value in update_values.items() if value is not None} # remove value which is None

        # read csv
        updated_rows = []
        with open(self.file_path, 'r', newline='') as file:
            reader = csv.DictReader(file)
            headers = reader.fieldnames

            for row in reader:
                if row.get('user_id') == str(user_id) or row.get('account') == account:
                    for key, new_value in update_values.items():
                        row[key] = new_value
                updated_rows.append(row)

        # write to csv
        with open(self.file_path, 'w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=headers)
            writer.writeheader()
            writer.writerows(updated_rows)

class Counseling_Record:
    def __init__(self, psychol_file_name: str = None, patient_file_name: str = None):
        self.psychol_file_name = "cnsl_info_psychol.xlsx" if psychol_file_name is None else psychol_file_name
        self.psychol_path = gen_file_path(self.psychol_file_name, 'data', 'counsel')
        self.patient_file_name = "cnsl_info_patient.xlsx" if patient_file_name is None else patient_file_name
        self.patient_path = gen_file_path(self.patient_file_name, 'data', 'counsel')
        self.testing = True if (psychol_file_name and patient_file_name) is not None else False
        self.glb_cfg = Config(False) if psychol_file_name is None or patient_file_name is None else Config(True)
        self.psychol_setting = self.glb_cfg.get_section('cnsl_info_psychol')
        self.psychol_topic = list(self.psychol_setting.keys())
        self._build_file(self.psychol_path, "Psycholist Info", self.psychol_topic)
        self.patient_setting = self.glb_cfg.get_section('cnsl_info_patient')
        self.patient_topic = list(self.patient_setting.keys())
        self._build_file(self.patient_path, "Patient Info", self.patient_topic)
        self.cnsl_rcrd_info_pat_setting = self.glb_cfg.get_section('cnsl_rcrd_info_pat')
        self.cnsl_rcrd_pat_setting = self.glb_cfg.get_section('cnsl_rcrd_pat')
        self.cnsl_rcrd_topic = list(self.cnsl_rcrd_pat_setting.keys())
        self.cnsl_rcrd_item = []
        self.cnsl_rcrd_item_size = []
        self.cnsl_rcrd_item_header = []
        for sub_topic in self.cnsl_rcrd_topic:
            cur_items = self.cnsl_rcrd_pat_setting[sub_topic]
            self.cnsl_rcrd_item.append(list(cur_items))
            self.cnsl_rcrd_item_size.append(len(cur_items))
            for sub_item in cur_items:
                self.cnsl_rcrd_item_header.append(sub_item)

    def _build_file(self, file_path, title, topic):
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = title

            # title
            headers = []
            for sub_topic in topic:
                headers.append(sub_topic)
            sheet.append(headers)

            workbook.save(file_path)

    def _build_rcrd(self, file_path):
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Counseling Data'

            # first title
            headers = []
            for cnsl_rcrd_topic, size in zip(self.cnsl_rcrd_topic, self.cnsl_rcrd_item_size):
                headers.append(cnsl_rcrd_topic)
                headers.extend([None] * (size - 1))
            sheet.append(headers)

            # second title
            headers = self.cnsl_rcrd_item_header
            sheet.append(headers)

            workbook.save(file_path)

    def _refresh_file(self): # only for unit-testing
        file_path = [self.psychol_path, self.patient_path]
        title = ['Psycholist Info', 'Patient Info']
        topic = [self.psychol_topic, self.patient_topic]
        for p, tt, tp in zip(file_path, title, topic):
            if os.path.exists(p):
                os.remove(p)
                self._build_file(p, tt, tp)

    def _get_cnsl_rcrd_info_pat_id(self):
        pat_id = self.cnsl_rcrd_info_pat_setting['nxt_patitent_id']
        max_pat_id = self.cnsl_rcrd_info_pat_setting['max_patitent_id']
        if pat_id > max_pat_id:
            return None
        self.glb_cfg.update_setting('cnsl_rcrd_info_pat', 'nxt_patitent_id', (pat_id + 1))
        self.glb_cfg.save_settings()
        return pat_id

    def _gen_cnsl_rcrd_file_name(self, id_num):
        return (str(id_num) + ".xlsx") if self.testing is False else (str(id_num) + "_test.xlsx")

    def set_cnsl_info_psychol(self, user_id, patient_id):
        check_type(user_id, int)
        check_type(patient_id, int)
        workbook = load_workbook(self.psychol_path)
        sheet = workbook["Psycholist Info"]
        row = [str(user_id), str(patient_id)]
        sheet.append(row)
        workbook.save(self.psychol_path)
        workbook.close()

    def get_cnsl_info_psychol(self, user_id):
        workbook = load_workbook(self.psychol_path)
        sheet = workbook["Psycholist Info"]
        patitent_id = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Starting from row 2 to exclude headers
            if row[0] == str(user_id):
                 patitent_id.append(int(row[1]))
        workbook.close()
        return patitent_id

    def set_cnsl_info_patient(self, name, id_num):
        check_type(name, str)
        check_type(id_num, int)
        workbook = load_workbook(self.patient_path)
        sheet = workbook["Patient Info"]
        row = [str(id_num), name]
        sheet.append(row)
        workbook.save(self.patient_path)
        workbook.close()

    def get_cnsl_info_patient(self, id_num):
        workbook = load_workbook(self.patient_path)
        sheet = workbook["Patient Info"]
        patitent_name = "No this patitent found"
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Starting from row 2 to exclude headers
            if row[0] == str(id_num):
                 patitent_name = row[1]
        workbook.close()
        return patitent_name

    def add_record(self, record, user_id = None, id_num = None):
        reset_file = False
        if id_num is None:
            if user_id is None: # no user_id to set
                return None
            reset_file = True
            id_num = self._get_cnsl_rcrd_info_pat_id()
            name = record['name']
            self.set_cnsl_info_patient(name, id_num)
            if id_num == None: # can't get id_num
                return None
            self.set_cnsl_info_psychol(user_id, id_num)
        file_path = gen_file_path(self._gen_cnsl_rcrd_file_name(id_num), 'data', 'counsel')
        if not os.path.exists(file_path):
            self._build_rcrd(file_path)

        workbook = load_workbook(file_path)
        sheet = workbook["Counseling Data"]
        if self.testing is True and reset_file is True:
            start_row = 2
            sheet.delete_rows(start_row + 1, sheet.max_row - start_row)
        row = [record.get(header, None) for header in self.cnsl_rcrd_item_header]
        sheet.append(row)
        workbook.save(file_path)
        workbook.close()
        return id_num

    def read_record(self, id_num, date = None):
        file_path = gen_file_path(self._gen_cnsl_rcrd_file_name(id_num), 'data', 'counsel')
        if not os.path.exists(file_path):
            raise FileNotFoundError("The file does not exist.")

        workbook = load_workbook(file_path)
        sheet = workbook["Counseling Data"]

        # Read headers (second row) and extract values
        header2 = [cell.value for cell in sheet[2]]

        # Read records and find the row with the matching id_num
        all_records = []
        for row in sheet.iter_rows(min_row=3, values_only=True):  # Starting from row 3 (where records start)
            records = {}
            for header, value in zip(header2, row):
                records[header] = value
            all_records.append(records)

        # find specific date records
        for records in all_records:
            for key in records:
                value = records[key]
                if value == date:
                    return records

        workbook.close()
        return all_records

def gen_file_path(file_name, sub_path1, sub_path2: str = None):
    cur_path = os.getcwd()
    sub_paths = [sub_path1]
    if sub_path2:
        sub_paths.append(sub_path2)
    file_path = os.path.join(cur_path, *sub_paths, file_name)
    return file_path

acnt_info = Account_Info()
cnsl_rcrd = Counseling_Record()